#!/usr/bin/python
# -*- coding: utf-8 -*-
import os
import sys
import importlib
import logging
import time
import datetime
from itertools import product
import types
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import range_boundaries
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl.drawing.image import Image

# def gbk2utf(in_data , tag):
# 	if 1 == tag:
# 		return in_data.encode('gbk').decode('gbk')
# 	elif 0 == tag:
# 		return in_data.encode('gbk').decode('gbk').encode('utf8')
#     elif 2 == tag:
#         return in_data.encode('utf8')        
def style_range(ws, cell_range, border=Border(), fill=None, font=None, alignment=None):
    """
    Apply styles to a range of cells as if they were a single cell.

    :param ws:  Excel worksheet instance
    :param range: An excel range to style (e.g. A1:F20)
    :param border: An openpyxl Border
    :param fill: An openpyxl PatternFill or GradientFill
    :param font: An openpyxl Font object
    """

    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)

    first_cell = ws[cell_range.split(":")[0]]
    if alignment:
        ws.merge_cells(cell_range)
        first_cell.alignment = alignment

    rows = ws[cell_range]
    if font:
        first_cell.font = font

    for cell in rows[0]:
        cell.border = cell.border + top
    for cell in rows[-1]:
        cell.border = cell.border + bottom

    for row in rows:
        l = row[0]
        r = row[-1]
        l.border = l.border + left
        r.border = r.border + right
        if fill:
            for c in row:
                c.fill = fill


def merge_cells(self, range_string=None, start_row=None, start_column=None, end_row=None, end_column=None):
    """ Set merge on a cell range.  Range is a cell range (e.g. A1:E1)
    This is monkeypatched to remove cell deletion bug
    https://bitbucket.org/openpyxl/openpyxl/issues/365/styling-merged-cells-isnt-working
    """
    if not range_string and not all((start_row, start_column, end_row, end_column)):
        msg = "You have to provide a value either for 'coordinate' or for\
        'start_row', 'start_column', 'end_row' *and* 'end_column'"
        raise ValueError(msg)
    elif not range_string:
        range_string = '%s%s:%s%s' % (get_column_letter(start_column),
                                        start_row,
                                        get_column_letter(end_column),
                                        end_row)
    elif ":" not in range_string:
        if COORD_RE.match(range_string):
            return  # Single cell, do nothing
        raise ValueError("Range must be a cell range (e.g. A1:E1)")
    else:
        range_string = range_string.replace('$', '')

    if range_string not in self._merged_cells:
        self._merged_cells.append(range_string)

def convert(config):

    # 1. 读取数据源excel的数据，读入内存
    path_source = os.path.join(os.getcwd(), config['read_params']['source_name'])
    doc_source = load_workbook(path_source)

    # 1.1 默认第一个工作簿(sheet_list_s[0])，输出表格行列数
    sheet_list_s = doc_source.sheetnames
    s = doc_source[sheet_list_s[0]]
    print("row and column number of data source: {0} | {1}".format(s.max_row, s.max_column))

    # 2. 读取模板excel的数据，读入内存
    path_template = os.path.join(os.getcwd(), config['read_params']['template_name'])
    doc_template = load_workbook(path_template, keep_vba=True)

    # 2.1 获取模板excel所有工作簿(sheet)
    sheet_list_t = doc_template.sheetnames
    print("names of modes: {}".format(sheet_list_t))
    print("num of modes: {}".format(len(sheet_list_t)))

    # 2.2 尝试读取第一个工作簿(sheet_list_t[0])，输出表格行列数
    temp0 = doc_template[sheet_list_t[0]]           # or s = doc_source.sheet_by_index(1)
    print("row and column number of template {0}: {1} | {2}".format(sheet_list_t[0], temp0.max_row, temp0.max_column))

    # 3. 建立子目录，用于生成excel文档
    sub_working_dir = '{}/{}/{}'.format(
        os.getcwd(), config['output_dir'],
        time.strftime("%d%H%M%S", time.localtime()))
    if not os.path.exists(sub_working_dir):
        os.makedirs(sub_working_dir)
    logging.info("sub working dir: %s" % sub_working_dir)

    # 4. 源数据每一行都进行处理
    for i in range(s.max_row - 2):
        row_index = i + 3

        workbook1 = load_workbook(path_template, keep_vba=True)                         # 建立模板sheet副本用于修改添加信息

        # 4.1 读取源数据中的测量范围，提取数字(上限，下限)
        up_lim = s.cell(column=7, row=row_index).value
        down_lim = s.cell(column=6, row=row_index).value
        print('up_lim: {}'.format(up_lim))

        # 4.2 根据上限判断使用哪个模板，建立工作簿副本w_sheet
        if up_lim == 0.16 or up_lim == 0.25 or up_lim == 0.4 or up_lim == 0.6 or up_lim == 1:
            sheet_index = 0                                                             # mode 1
            print
        elif up_lim == 1.6 or up_lim == 2.5 or up_lim == 4 or up_lim == 6 or up_lim == 10:
            sheet_index = 1                                                             # mode 2
        elif up_lim == 16 or up_lim == 25 or up_lim == 40 or up_lim == 60 or up_lim == 100:
            sheet_index = 2                                                             # mode 3
        elif up_lim == 160 or up_lim == 250:
            sheet_index = 3                                                             # mode 4
        else:
            print("Other type!")

        # 4.3 副本去掉其他无用的工作簿sheet
        for sh in range(len(sheet_list_t)):
            if sh == sheet_index:
                continue
            workbook1.remove(workbook1.get_sheet_by_name(sheet_list_t[sh]) )   

        # 5. 确定worksheet为待修改的工作簿sheet
        sheet_list_0 = workbook1.sheetnames
        worksheet = workbook1[sheet_list_0[0]]

        # 5.1 送检单位
        worksheet['S4'] = s.cell(column=1, row=row_index).value
        # 5.2 流水号
        worksheet['S5'] = s.cell(column=2, row=row_index).value
        # 5.3 计量器具名称
        worksheet['N4'] = s.cell(column=3, row=row_index).value
        # 5.4 制造单位
        worksheet['S6'] = s.cell(column=4, row=row_index).value
        # 5.5 检定地点
        worksheet['L5'] = s.cell(column=5, row=row_index).value
        # 5.6 测量范围
        worksheet['T7'] = down_lim
        worksheet['V7'] = up_lim
        # 5.7 计量单位
        worksheet['X7'] = s.cell(column=8, row=row_index).value
        # 5.8 出厂编号
        worksheet['S9'] = s.cell(column=9, row=row_index).value
        # 5.9 精度等级
        worksheet['S8'] = s.cell(column=10, row=row_index).value
        # 5.10 温度
        worksheet['J9'] = s.cell(column=11, row=row_index).value
        # 5.11 湿度
        worksheet['N9'] = s.cell(column=12, row=row_index).value
        # 5.12 检定日期（年月日）
        worksheet['S12'] = s.cell(column=13, row=row_index).value
        worksheet['U12'] = s.cell(column=14, row=row_index).value
        worksheet['W12'] = s.cell(column=15, row=row_index).value

        worksheet.merge_cells('B3:H3')
        
        thin = Side(border_style="thin", color="000000")
        thick = Side(border_style="medium", color="000000")

        border_top = Border(top=thick, left=thin, right=thick, bottom=thin)
        border_normal = Border(top=thin, left=thin, right=thick, bottom=thin)
        border_bottom = Border(top=thin, left=thin, right=thick, bottom=thick)
        border_right = Border(top=thin, left=thin, right=thin, bottom=thin)
        border_right_thick = Border(top=thin, left=thin, right=thick, bottom=thin)

        style_range(worksheet, 'B3:P3', border=border_top)
        style_range(worksheet, 'B10:P10', border=border_top)
        style_range(worksheet, 'B18:P18', border=border_top)

        style_range(worksheet, 'B4:P4', border=border_normal)
        style_range(worksheet, 'B5:P5', border=border_normal)
        style_range(worksheet, 'B6:P6', border=border_normal)
        style_range(worksheet, 'B7:P7', border=border_normal)
        style_range(worksheet, 'B8:P8', border=border_normal)
        style_range(worksheet, 'B11:P11', border=border_normal)
        style_range(worksheet, 'B12:P12', border=border_normal)
        style_range(worksheet, 'B13:P13', border=border_normal)
        style_range(worksheet, 'B14:P14', border=border_normal)
        style_range(worksheet, 'B15:P15', border=border_normal)
        style_range(worksheet, 'B16:P16', border=border_normal)
        style_range(worksheet, 'B19:P19', border=border_normal)
        style_range(worksheet, 'B20:P20', border=border_normal)
        # style_range(worksheet, 'B21:P21', border=border_normal)
        style_range(worksheet, 'B22:P22', border=border_normal)
        style_range(worksheet, 'B23:P23', border=border_normal)
        style_range(worksheet, 'B24:P24', border=border_normal)
        style_range(worksheet, 'B25:P25', border=border_normal)
        style_range(worksheet, 'B26:P26', border=border_normal)
        style_range(worksheet, 'B27:P27', border=border_normal)
        style_range(worksheet, 'B28:P28', border=border_normal)
        style_range(worksheet, 'B29:P29', border=border_normal)
        # style_range(worksheet, 'B30:P30', border=border_normal)
        style_range(worksheet, 'B31:P31', border=border_normal)
        style_range(worksheet, 'B32:P32', border=border_normal)
        style_range(worksheet, 'B33:P33', border=border_normal)
        style_range(worksheet, 'B34:P34', border=border_normal)
        style_range(worksheet, 'B35:P35', border=border_normal)
        style_range(worksheet, 'B36:P36', border=border_normal)

        style_range(worksheet, 'B9:P9', border=border_bottom)
        style_range(worksheet, 'B17:P17', border=border_bottom)
        style_range(worksheet, 'B37:P37', border=border_bottom)

        style_range(worksheet, 'B21:L21', border=border_right)
        style_range(worksheet, 'M21:N21', border=border_right)
        style_range(worksheet, 'O21:P21', border=border_right_thick)
        style_range(worksheet, 'B30:L30', border=border_right)
        style_range(worksheet, 'M30:N30', border=border_right)
        style_range(worksheet, 'O30:P30', border=border_right_thick)

        img = Image(config['read_params']['picture_name'], size=(600, 480))
        worksheet.add_image(img, 'A31')

        # stem, ext = os.path.splitext(path_template)
        path_write = os.path.join(sub_working_dir, config['output_params']['output_name'] + str(i + 1) + '.xlsm')
        # path_write = path_write.format(stem, __version__)
        workbook1.save(filename = path_write)


def main():
    logging.basicConfig(level=logging.DEBUG,
                        format="[%(asctime)s %(filename)s] %(message)s")

    if len(sys.argv) != 2:
        logging.error("Usage: python training.py params.py")
        sys.exit()
    params_path = sys.argv[1]
    if not os.path.isfile(params_path):
        logging.error("no params file found! path: {}".format(params_path))
        sys.exit()
    config = importlib.import_module(params_path[:-3]).PARAMS
    
    convert(config)

if __name__ == "__main__":
    main()